import openpyxl as xl
import os

from openpyxl.styles import PatternFill
from pathlib import Path

path = Path("sheets")

xlsx = ''

try:
    def remove_fill(ws):
        # ws is not the worksheet name, but the worksheet object
        no_fill = PatternFill(fill_type=None)
        # for row2 in ws.iter_rows():
        for row2 in ws['A1':'Z150']:
            for cell2 in row2:
                cell2.fill = no_fill


    print('Loading files ...')
    for root, dirs, files in os.walk(path):
        for file in files:
            if not file.endswith(".xlsx"):
                continue
            xlsx += os.path.join(root, file) + ','

    paths = xlsx.split(',')
    paths.__delitem__(-1)

    if len(paths) > 0:
        print('Starting clean-up ...')
    else:
        print('No files were found ...')

    for i, file in enumerate(paths):
        print(f'Loading file ({i+1} / {len(paths)}): {file}')
        wb = xl.load_workbook(file)
        sheet = wb.worksheets[0]
        print(f'Worksheet: {sheet.title}')

        # delete formulas
        for row in sheet['A1':'G100']:
            for cell in row:
                cell.value = None
        print('Formulas deleted ... clearing fill ...')

        # clear coloring
        remove_fill(sheet)
        print('Done ...')

        # save the updated xl file
        print(f'Saving file {file} ...')
        wb.save(file)

    print(f'{len(paths)} files were processed.')
except ValueError:
    print('Error detected.')
except AttributeError:
    print('Attribute value is read-only.')
